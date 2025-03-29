import openpyxl
from openpyxl import load_workbook
import matplotlib
matplotlib.use('TkAgg')
import matplotlib.pyplot as plt
import numpy as np
import pickle
from sklearn.model_selection import train_test_split
from sklearn.metrics import accuracy_score, recall_score, f1_score, confusion_matrix, ConfusionMatrixDisplay
from sklearn.tree import DecisionTreeClassifier
from sklearn.naive_bayes import GaussianNB
from sklearn.neighbors import KNeighborsClassifier
from sklearn.neural_network import MLPClassifier
from sklearn.ensemble import RandomForestClassifier
from sklearn.svm import SVC


# DecisionTree
def net(path):

    workbook = load_workbook(path)
    sheet = workbook.active

    x_values = []
    y_values = []
    for row in sheet.iter_rows():
        row_data = [cell.value for cell in row]

        x = row_data[1:-1]
        y = row_data[len(row_data) - 1]

        x_values.append(x)
        y_values.append(y)

    features = x_values[0]
    x_values = np.array(x_values[1:])
    y_values = np.array(y_values[1:])

    #P = compare4(x_values, y_values, features)
    #x_values = np.array(bian(x_values, P))

    x_train, x_test, y_train, y_test = train_test_split(x_values, y_values, test_size=0.2, random_state=42,shuffle=True)

    model1 = DecisionTreeClassifier(max_depth=None, min_samples_split=2, random_state=42)
    model2 = GaussianNB()
    model3 = KNeighborsClassifier(n_neighbors=5, weights='uniform', algorithm='auto', metric='euclidean')
    model4 = MLPClassifier(hidden_layer_sizes=(100,), max_iter=300, alpha=1e-4, solver='sgd', verbose=10, random_state=1,learning_rate_init=.1)
    model5 = RandomForestClassifier(n_estimators=100, max_depth=None, random_state=42)
    model6 = SVC(kernel='linear', C=1)

    model=model1


    model.fit(x_train, y_train)


    U=[]

    y_pred = model.predict(x_test)
    #print(y_pred)

    accuracy = accuracy_score(y_test, y_pred)
    print(f"Accuracy: {accuracy:.3f}")
    U.append(round(accuracy, 3))

    # 计算并输出召回率
    recall = recall_score(y_test, y_pred, average='weighted')
    print(f"Recall: {recall:.3f}")
    #U.append(round(recall, 3))

    # 计算特异性（Specificity），需要手动计算每个类别
    conf_matrix = confusion_matrix(y_test, y_pred)
    num_classes = conf_matrix.shape[0]
    specificity_list = []

    for i in range(num_classes):
        tn = np.sum(np.delete(np.delete(conf_matrix, i, axis=0), i, axis=1))
        fp = np.sum(np.delete(conf_matrix[:, i], i))
        fn = np.sum(np.delete(conf_matrix[i, :], i))
        tp = conf_matrix[i, i]

        specificity = tn / (tn + fp) if (tn + fp) > 0 else 0
        specificity_list.append(specificity)

    average_specificity = np.mean(specificity_list)
    print(f"Average Specificity: {average_specificity:.3f}")
    #U.append(round(average_specificity, 3))


    # 计算并输出F1分数
    f1 = f1_score(y_test, y_pred, average='weighted')
    print(f"F1 Score: {f1:.3f}")
    #U.append(round(f1, 3))


    averagevalue=0
    for i in range(1, 5):
        print(i, "：")
        A = []
        B = []
        for j in range(0, len(y_test)):
            if y_test[j] == i:
                A.append(i)
                B.append(x_test[j])
        #print(A)

        y_pred = model.predict(B)
        #print(y_pred)

        accuracy = accuracy_score(A, y_pred)
        averagevalue=averagevalue+accuracy
        print(f"Accuracy: {accuracy:.3f}")
        U.append(round(accuracy, 3))

        # 计算并输出召回率
        recall = recall_score(A, y_pred, average='weighted')
        print(f"Recall: {recall:.3f}")

        # 计算特异性（Specificity），需要手动计算每个类别
        conf_matrix = confusion_matrix(A, y_pred)
        num_classes = conf_matrix.shape[0]
        specificity_list = []

        for i in range(num_classes):
            tn = np.sum(np.delete(np.delete(conf_matrix, i, axis=0), i, axis=1))
            fp = np.sum(np.delete(conf_matrix[:, i], i))
            fn = np.sum(np.delete(conf_matrix[i, :], i))
            tp = conf_matrix[i, i]

            specificity = tn / (tn + fp) if (tn + fp) > 0 else 0
            specificity_list.append(specificity)

        average_specificity = np.mean(specificity_list)
        print(f"Average Specificity: {average_specificity:.3f}")

        # 计算并输出F1分数
        f1 = f1_score(A, y_pred, average='weighted')
        print(f"F1 Score: {f1:.3f}")

    print(f"Average Value: {averagevalue/4:.3f}")

    U.append(round(averagevalue/4, 3))

    model_path = './output/model.pkl'
    pickle.dump(model, open(model_path, 'wb'))

    return U

# L1-based feature selection
def compare4(x,y,feature_names):
    from sklearn.linear_model import Lasso
    from sklearn.preprocessing import StandardScaler
    scaler = StandardScaler()

    lasso = Lasso(alpha=0.2)
    lasso.fit(x, y)
    lasso.coef_
    #print("Selected features4:", lasso.coef_)

    P = lasso.coef_
    return P


def bian(X,P):
    new_x =[]
    for i in range(0,len(X)):
        x=[]
        for j in range(0, len(P)):
            x.append(X[i][j]*P[j])
        new_x.append(x)
    return new_x


if __name__ == '__main__':

    path1 = './Excel/heart.xlsx'
    path2 = './Excel/add-cot.xlsx'
    path3 = './Excel/add-video.xlsx'
    path4 = './Excel/add.xlsx'
    path8 = './Compare1/Excel2/compare16.xlsx'
    path9 = './Excel/Gate_Attention_LMF.xlsx'
    path=path9

    #net(path)

    wb = openpyxl.Workbook()  # 创建一个excel文件
    sheet1 = wb.active  # 获得一个的工作表

    sheet1.append(net(path))
    wb.save('./all.xlsx')