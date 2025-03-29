import matplotlib

import Gate

matplotlib.use('TkAgg')
import matplotlib.pyplot as plt
import random
import cv2
import mediapipe as mp
import numpy as np
import os
import openpyxl
from openpyxl import load_workbook
import math
import torch
import torch.nn as nn
import torch.nn.functional as F
import matplotlib.pyplot as plt
from PIL import Image
import torchvision.transforms as transforms
import Gate
import Attention
import LMF
import openpyxl
from openpyxl import load_workbook
import matplotlib
matplotlib.use('TkAgg')
import matplotlib.pyplot as plt
import numpy as np
import pickle
from sklearn.model_selection import train_test_split
from sklearn.metrics import accuracy_score, recall_score, f1_score, confusion_matrix, ConfusionMatrixDisplay
from sklearn.tree import DecisionTreeClassifier
from sklearn.naive_bayes import GaussianNB
from sklearn.neighbors import KNeighborsClassifier
from sklearn.neural_network import MLPClassifier
from sklearn.ensemble import RandomForestClassifier
from sklearn.svm import SVC
import numpy as np
import pandas as pd
from sklearn.model_selection import train_test_split
from sklearn.preprocessing import StandardScaler
from sklearn.metrics import precision_score, recall_score, f1_score, confusion_matrix
from tensorflow.keras.models import Sequential
from tensorflow.keras.layers import Dense
from tensorflow.keras.utils import to_categorical


mpHands = mp.solutions.hands
hands = mpHands.Hands(static_image_mode=False,
                      max_num_hands=1,
                      min_detection_confidence=0.75,
                      min_tracking_confidence=0.5
                      )
mpDraw = mp.solutions.drawing_utils

# 创建一个卷积层（示例卷积核）
conv_layer1 = nn.Conv2d(in_channels=3, out_channels=1, kernel_size=1, padding=1)
conv_layer3 = nn.Conv2d(in_channels=3, out_channels=1, kernel_size=3, padding=1)
conv_layer5 = nn.Conv2d(in_channels=3, out_channels=1, kernel_size=5, padding=1)

# 最大池化层
max_pool = nn.MaxPool2d(kernel_size=2, stride=2)
# 平均池化层
avg_pool = nn.AvgPool2d(kernel_size=2, stride=2)

# 激活函数
relu = nn.ReLU()
sigmoid = nn.Sigmoid()


#随机选取
def lottery_draw(n, start, end):
    numbers = list(range(start, end + 1))
    lucky_numbers = random.sample(numbers, n)
    return lucky_numbers

#读取execl
def readexecl(path):
    workbook = load_workbook(path)
    sheet = workbook.active

    D = []
    n = sheet.max_column
    for i in range(0,n):
        A=[]
        D.append(A)

    for row in sheet.iter_rows():

        row_data = [cell.value for cell in row]

        for i in range(0, n):

            #D[i].append(float(row_data[i] - row_data[i % 3]))
            D[i].append(float(row_data[i]))
    return D



def C1(img):

    output_conv = conv_layer1(img)
    #print(output_conv)
    output_pool =max_pool(output_conv)
    #print(output_pool)
    flatten_output = torch.flatten(output_pool, 1)

    fc_layer = nn.Linear(len(flatten_output[0]), 16)
    x = fc_layer(flatten_output[0])
    x = x.tolist()

    return x


def C2(img):
    x1 = avg_pool(img)
    x2 = torch.flatten(x1, 1)

    fc_layer = nn.Linear(len(x2[0]), 128)
    x3 = fc_layer(x2)

    x4 = relu(x3)

    fc_layer = nn.Linear(len(x4[0]), 16)
    x5 = fc_layer(x4)

    x6 = sigmoid(x5)

    x6 = x6.tolist()

    return x6[0]

def imgturn(image):

    pil_image = Image.fromarray(image)
    transform = transforms.Compose([transforms.Resize((224, 224)),
                                    transforms.ToTensor()])
    input_image = transform(pil_image).unsqueeze(0)  # 添加一个批次维度
    # print(input_image)
    return input_image


def picdata(datapath):


            img = cv2.imread(datapath)

            E=[]


            imgRGB = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)
            results = hands.process(imgRGB)

            if results.multi_hand_landmarks:

                img2=imgturn(imgRGB)
                A=C1(img2)+C2(img2)

                for hand_landmarks in results.multi_hand_landmarks:
                    #print('hand_landmarks:',hand_landmarks)

                    B = []

                    x0 = 0
                    y0 = 0
                    z0 = 0
                    x1 = 0
                    y1 = 0
                    z1 = 0

                    for id, lm in enumerate(hand_landmarks.landmark):

                        #print("n={}, x = {}, y = {}, z = {}".format(id, lm.x, lm.y, lm.z))
                        if id == 0:
                            x0 = lm.x
                            y0 = lm.y
                            z0 = lm.z
                        if id == 1:
                            x1 = lm.x
                            y1 = lm.y
                            z1 = lm.z
                        if id<=20 and id>=0:
                            k=math.sqrt((x1 - x0) * (x1 - x0) + (y1 - y0) * (y1 - y0) + (z1 - z0) * (z1 - z0))
                            B.append((lm.x-x0)/k)
                            B.append((lm.y-y0)/k)
                            B.append((lm.z-z0)/k)

                    V0x = B[0]
                    V0y = B[1]
                    V0z = B[2]

                    V4x = B[12]
                    V4y = B[13]
                    V4z = B[14]

                    V5x = B[15]
                    V5y = B[16]
                    V5z = B[17]

                    V6x = B[18]
                    V6y = B[19]
                    V6z = B[20]

                    V8x = B[24]
                    V8y = B[25]
                    V8z = B[26]

                    X1t = V4x - V0x
                    X2t = V8x - V0x

                    Y1t = V4y - V0y
                    Y2t = V8y - V0y

                    I1 = [V4x - V0x, V4y - V0x, V4z - V0z]
                    J1 = [V8x - V0x, V8y - V0x, V8z - V0z]

                    I2 = [V5x - V6x, V5y - V6x, V5z - V6z]
                    J2 = [V8x - V6x, V8y - V6x, V8z - V6z]

                    m1 = I1[0] * J1[0] + I1[1] * J1[1] + I1[2] * J1[2]

                    m2 = I1[0] * I1[0] + I1[1] * I1[1] + I1[2] * I1[2]

                    O1 = math.acos((I1[0] * J1[0] + I1[1] * J1[1] + I1[2] * J1[2]) / (
                                math.sqrt(I1[0] * I1[0] + I1[1] * I1[1] + I1[2] * I1[2]) * math.sqrt(
                            J1[0] * J1[0] + J1[1] * J1[1] + J1[2] * J1[2])))
                    O2 = math.acos((I2[0] * J2[0] + I2[1] * J2[1] + I2[2] * J2[2]) / (
                                math.sqrt(I2[0] * I2[0] + I2[1] * I2[1] + I2[2] * I2[2]) * math.sqrt(
                            J2[0] * J2[0] + J2[1] * J2[1] + J2[2] * J2[2])))

                    C = [X1t, X2t, Y1t, Y2t, O1, O2]


                    return A,C

            return 0,0






def cot_feture3(V5X, V5Y, V5Z,V7X, V7Y, V7Z):
    dx = []
    for i in range(0, len(V5X)):
        dx.append(V5X[i] - V7X[i])
    dy = []
    for i in range(0, len(V5X)):
        dy.append(V5Y[i] - V7Y[i])
    dz = []
    for i in range(0, len(V5X)):
        dz.append(V5Z[i] - V7Z[i])

    d=[]
    for i in range(0, len(V5X)):
        d.append(dx[i]*dx[i]+dy[i]*dy[i]+dz[i]*dz[i])

    return  max(d), min(d)




def getdata(path):
    X=[]
    Y=[]

    for i in range(0,4):
        path1=path+str(i)+'/'
        print(path1)

        C = []
        D = []

        file0 = os.listdir(path1)
        for f0 in file0:
            # 字符串拼接
            datapath = path1 + f0

            A,B=picdata(datapath)
            if A!=0:
                #print(datapath)
                #print(A,B)
                C.append(A)
                D.append(B)
                Y.append(i)

        M = np.hstack((C, D))
        W=Attention.attention(M,M)

        #print(W)

        X.extend(W)


    x_values = np.array(X)
    y_values = np.array(Y)



    x_train, x_test, y_train, y_test = train_test_split(x_values, y_values, test_size=0.2, random_state=42,shuffle=True)

    scaler = StandardScaler()
    x_train = scaler.fit_transform(x_train)
    x_test = scaler.transform(x_test)


    # 将标签转换为one-hot编码
    y_train = to_categorical(y_train, num_classes=4)

    # 构建DNN模型
    model = Sequential()
    model.add(Dense(64, activation='relu', input_shape=(x_train.shape[1],)))
    model.add(Dense(32, activation='relu'))
    model.add(Dense(4, activation='softmax'))  # 使用softmax激活函数进行四分类

    # 编译模型
    model.compile(optimizer='adam', loss='categorical_crossentropy', metrics=['accuracy'])

    # 训练模型
    model.fit(x_train, y_train, epochs=50, batch_size=32, validation_split=0.2)

    # 在测试集上进行预测
    y_pred = model.predict(x_test)

    y_pred_classes = np.argmax(y_pred, axis=1)
    print(y_pred_classes)

    # 计算评估指标
    precision = precision_score(y_test, y_pred_classes, average='weighted')
    recall = recall_score(y_test, y_pred_classes, average='weighted')
    f1 = f1_score(y_test, y_pred_classes, average='weighted')

    # 计算Specificity
    cm = confusion_matrix(y_test, y_pred_classes)
    fp = cm.sum(axis=0) - np.diag(cm)
    fn = cm.sum(axis=1) - np.diag(cm)
    tp = np.diag(cm)
    tn = cm.sum() - (fp + fn + tp)

    specificity = tn / (tn + fp)
    specificity = np.mean(specificity)

    # 打印评估结果
    print(f'Precision: {precision:.4f}')
    print(f'Recall: {recall:.4f}')
    print(f'Specificity: {specificity:.4f}')
    print(f'F1 Score: {f1:.4f}')


if __name__ == '__main__':

    path = './Pic3/'

    getdata(path)
