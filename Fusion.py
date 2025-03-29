import matplotlib
matplotlib.use('TkAgg')
import random
import cv2
import mediapipe as mp
import numpy as np
import os
import openpyxl
from openpyxl import load_workbook
import math
import torch
import torch.nn as nn
import torch.nn.functional as F
import matplotlib.pyplot as plt
from PIL import Image
import torchvision.transforms as transforms
import random
import Gate
import Attention
import LMF
import MFB
import MCB

def Fusion(path1,path2,savepath):

    wb = openpyxl.Workbook()  # 创建一个excel文件
    sheet1 = wb.active  # 获得一个的工作表

    F = []
    F.append('Num')
    for i in range(1, 10):
        F.append(i)
    F.append('Type')

    sheet1.append(F)

    n=0
    for i in [1,2,3,4]:

        M1 = []
        M2 = []
        M3 = []
        M4 = []

        for j in range(0,200):

            #n=n+1
            A=[]
            A.append(n)

            datapath1 = path1 + str(i) + '.xlsx'
            datapath2 = path2 + str(i) + '.xlsx'

            L1=select(datapath1)[1:-1]
            L2=select(datapath2)[1:-1]

            P1=L1[0:3]
            P2=L1[3:9]
            P3=L2[0:4]
            P4=L2[4:9]


            M1.append(P1)
            M2.append(P2)
            M3.append(P3)
            M4.append(P4)



        P1 = Gate.gate(M1, M2)
        P2 = Gate.gate(M2, M1)
        P3 = Gate.gate(M3, M4)
        P4 = Gate.gate(M4, M3)

        C = np.hstack((P1, P2))
        D = np.hstack((P3, P4))
        #print(C)

        C = Attention.attention(C, C)
        D = Attention.attention(D, D)
        #print(C,D)

        #E=MFB.Mfb(C,D)
        #E=MCB.Mcb(C,D)

        E=LMF.Lmf(C,D)


        for k in range(0,len(E)):
            F=[]

            n=n+1
            F.append(n)
            #print(E[k])
            W = [*F, *E[k]]

            W.append(i)
            #print(W)
            sheet1.append(W)


    wb.save(savepath)
    return

def select(datapath):

    workbook = load_workbook(datapath)
    sheet = workbook.active
    row_count = sheet.max_row

    random_number = random.randint(2, row_count)
    #print(random_number, row_count)

    i = 1
    for row in sheet.iter_rows():
        if i==random_number:
            row_data = [cell.value for cell in row]
            B = row_data
        i = i + 1
    return B

if __name__ == '__main__':

    path1 = './Excel/Cot/'
    path2 = './Excel/Video/'
    savepath = './Excel/fusion_this_study.xlsx'

    savepath2 = './Excel/add_Attention_LMF.xlsx'
    Fusion(path1,path2,savepath2)



    path3 = './Compare1/Excel2/Cot2/'
    path4 = './Compare1/Excel2/Video8/'
    savepath2 = './Compare1/Excel2/compare16.xlsx'
    #Fusion(path3, path4, savepath2)

