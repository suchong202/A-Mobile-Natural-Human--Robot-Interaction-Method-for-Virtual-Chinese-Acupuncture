import matplotlib

import Gate

matplotlib.use('TkAgg')
import matplotlib.pyplot as plt
import random
import cv2
import mediapipe as mp
import numpy as np
import os
import openpyxl
from openpyxl import load_workbook
import math
import torch
import torch.nn as nn
import torch.nn.functional as F
import matplotlib.pyplot as plt
from PIL import Image
import torchvision.transforms as transforms



mpHands = mp.solutions.hands
hands = mpHands.Hands(static_image_mode=False,
                      max_num_hands=1,
                      min_detection_confidence=0.75,
                      min_tracking_confidence=0.5
                      )
mpDraw = mp.solutions.drawing_utils

# 创建一个卷积层（示例卷积核）
conv_layer1 = nn.Conv2d(in_channels=3, out_channels=1, kernel_size=1, padding=1)
conv_layer3 = nn.Conv2d(in_channels=3, out_channels=1, kernel_size=3, padding=1)
conv_layer5 = nn.Conv2d(in_channels=3, out_channels=1, kernel_size=5, padding=1)

# 最大池化层
max_pool = nn.MaxPool2d(kernel_size=2, stride=2)
# 平均池化层
avg_pool = nn.AvgPool2d(kernel_size=2, stride=2)

# 激活函数
relu = nn.ReLU()
sigmoid = nn.Sigmoid()


#随机选取
def lottery_draw(n, start, end):
    numbers = list(range(start, end + 1))
    lucky_numbers = random.sample(numbers, n)
    return lucky_numbers

#读取execl
def readexecl(path):
    workbook = load_workbook(path)
    sheet = workbook.active

    D = []
    n = sheet.max_column
    for i in range(0,n):
        A=[]
        D.append(A)

    for row in sheet.iter_rows():

        row_data = [cell.value for cell in row]

        for i in range(0, n):

            #D[i].append(float(row_data[i] - row_data[i % 3]))
            D[i].append(float(row_data[i]))
    return D

def extend(r, n):
    num1 = int(n/len(r))
    num2 = int(n%len(r))

    L =[]

    if n>=len(r):
        for i in range(0, num1):
            for j in range(0, len(r)):
                L.append(r[j])

        for j in range(0, num2):
            L.append(r[j])

    else:
        for j in range(0, num2):
            L.append(r[j])

    return L

def cut(r, n):

   a = int((len(r) / 2) - (n*len(r))/2)
   b = int((len(r) / 2) + (n*len(r))/2)
   L = r[a:b]

   return L

def guiyi(r):
    min_val = min(r)
    max_val = max(r)
    normalized_arr = [(x - min_val) / (max_val - min_val) for x in r]

    return normalized_arr

def C1(img):

    output_conv = conv_layer1(img)
    #print(output_conv)
    output_pool =max_pool(output_conv)
    #print(output_pool)
    flatten_output = torch.flatten(output_pool, 1)

    fc_layer = nn.Linear(len(flatten_output[0]), 1)
    x = fc_layer(flatten_output[0])
    x = x.tolist()

    return x


def C2(img):
    x1 = avg_pool(img)
    x2 = torch.flatten(x1, 1)

    fc_layer = nn.Linear(len(x2[0]), 128)
    x3 = fc_layer(x2)

    x4 = relu(x3)

    fc_layer = nn.Linear(len(x4[0]), 4)
    x5 = fc_layer(x4)

    x6 = sigmoid(x5)

    x6 = x6.tolist()

    return x6[0]

def imgturn(image):

    pil_image = Image.fromarray(image)
    transform = transforms.Compose([transforms.Resize((224, 224)),
                                    transforms.ToTensor()])
    input_image = transform(pil_image).unsqueeze(0)  # 添加一个批次维度
    # print(input_image)
    return input_image


def getdata(datapath,sheet):

        cap = cv2.VideoCapture(datapath)
        success, frame = cap.read()
        if not success:
            print("无法读取视频")
        else:
            # 获取图像宽度和高度
            width = frame.shape[1]
            height = frame.shape[0]

        #统计帧数
        i = 0

        while True:

            sussess, img = cap.read()

            if not sussess:
                break

            imgRGB = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)
            results = hands.process(imgRGB)

            if results.multi_hand_landmarks:

                img2=imgturn(imgRGB)
                A=C1(img2)+C2(img2)

                for hand_landmarks in results.multi_hand_landmarks:
                    #print('hand_landmarks:',hand_landmarks)

                    B = []

                    x0 = 0
                    y0 = 0
                    z0 = 0
                    x1 = 0
                    y1 = 0
                    z1 = 0

                    for id, lm in enumerate(hand_landmarks.landmark):

                        #print("n={}, x = {}, y = {}, z = {}".format(id, lm.x, lm.y, lm.z))
                        if id == 0:
                            x0 = lm.x
                            y0 = lm.y
                            z0 = lm.z
                        if id == 1:
                            x1 = lm.x
                            y1 = lm.y
                            z1 = lm.z
                        if id<=8 and id>=1:
                            k=math.sqrt((x1 - x0) * (x1 - x0) + (y1 - y0) * (y1 - y0) + (z1 - z0) * (z1 - z0))
                            B.append((lm.x-x0)/k)
                            B.append((lm.y-y0)/k)
                            B.append((lm.z-z0)/k)
                    C=B+A
                    sheet.append(C)

                    # 创建一个500*500,3颜色通道图片的numpy矩阵
                    img = np.zeros((height,width,  3), dtype=np.uint8)
                    img[:] = (255, 255, 255)  # 白色背景

                    # 关键点可视化
                    handLmsStyle = mpDraw.DrawingSpec(color=(0, 0, 255), thickness=3)
                    handConStyle = mpDraw.DrawingSpec(color=(0, 255, 0), thickness=2)
                    mpDraw.draw_landmarks(img, hand_landmarks, mpHands.HAND_CONNECTIONS, handLmsStyle, handConStyle)

            #cv2.imshow('image', img)
            cv2.imwrite('./Node/' + str(i) + '.png', img)

            #清除之前线条
            img = np.zeros((width,height, 3), np.uint8)
            img.fill(0)
            i = i + 1

            if cv2.waitKey(1) & 0xFF == 27:
                break

        cap.release()

        return sheet

# 视频转excel文件
def videotoexecl(url, savepath):

    wb = openpyxl.Workbook()  # 创建一个excel文件
    sheet = wb.active  # 获得一个的工作表

    file0 = os.listdir(url)
    for f0 in file0:
        # 字符串拼接
        datapath = url + f0
        sheet=getdata(datapath,sheet)

    wb.save(savepath)


def Shi(V4X,V4Y,V4Z,V8X,V8Y,V8Z):
    flag=0
    A=[]
    B=[]
    C=[]
    n=0
    for j in range(0, len(V4X)):
        x = (V4X[j] - V8X[j])
        y = (V4Y[j] - V8Y[j])
        z = (V4Z[j] - V8Z[j])
        A.append(y)
    for j in range(0, len(A)):
        if A[j]<=0:
           n=n+1

    #print(n)
    if n>=10 and n<=40:
        flag=1

    return flag

def find_peaks(arr):
    peaks = [arr[0]]  # 初始峰值列表，包含第一个元素
    for i in range(1, len(arr) - 1):
        if arr[i] > arr[i - 1] and arr[i] > arr[i + 1]:
            peaks.append(arr[i])
    peaks.append(arr[-1])  # 添加最后一个元素，如果它是峰值
    return peaks

def cot_feture1(V4X, V4Y, V4Z,V7X, V7Y, V7Z,V8X, V8Y, V8Z):
    d1x =[]
    for i in range(0,len(V4X)):
        d1x.append(V4X[i] - V8X[i])
    d1y = []
    for i in range(0, len(V4Y)):
        d1y.append(V4Y[i] - V8Y[i])
    d1z = []
    for i in range(0, len(V4Z)):
        d1z.append(V4Z[i] - V8Z[i])

    d2x =[]
    for i in range(0,len(V4X)):
        d2x.append(V4X[i] - V7X[i])
    d2y = []
    for i in range(0, len(V4Y)):
        d2y.append(V4Y[i] - V7Y[i])
    d2z = []
    for i in range(0, len(V4Z)):
        d2z.append(V4Z[i] - V7Z[i])

    d1=[]
    d2=[]

    for i in range(0,len(d1x)):
        d1.append(d1x[i]*d1x[i]+d1y[i]*d1y[i]+d1z[i]*d1z[i])
    for i in range(0,len(d2x)):
        d2.append(d2x[i]*d2x[i]+d2y[i]*d2y[i]+d2z[i]*d2z[i])
    #print(d1,d2)

    return len(find_peaks(d1))+len(find_peaks(d2))

def cot_feture2(V6Y):
    max_value = max(V6Y)
    min_value = min(V6Y)
    #print(max_value - min_value)
    return max_value - min_value

def cot_feture3(V5X, V5Y, V5Z,V7X, V7Y, V7Z):
    dx = []
    for i in range(0, len(V5X)):
        dx.append(V5X[i] - V7X[i])
    dy = []
    for i in range(0, len(V5X)):
        dy.append(V5Y[i] - V7Y[i])
    dz = []
    for i in range(0, len(V5X)):
        dz.append(V5Z[i] - V7Z[i])

    d=[]
    for i in range(0, len(V5X)):
        d.append(dx[i]*dx[i]+dy[i]*dy[i]+dz[i]*dz[i])

    return  max(d), min(d)

def model_data(execlpath,savepath):

    wb = openpyxl.Workbook()  # 创建一个excel文件
    sheet = wb.active  # 获得一个的工作表

    name = ['Num']
    for i in range(1, 10):
        name.append(str(i))
    name.append('Type')

    sheet.append(name)


    D = readexecl(execlpath)

    L=10
    # len(D[0])-L

    n=0
    for j in range(0, len(D[0])-L+1):
        #print(j,j+10)

        V4X = D[9][j:j + L]
        V4Y = D[10][j:j + L]
        V4Z = D[11][j:j + L]
        V5X = D[12][j:j + L]
        V5Y = D[13][j:j + L]
        V5Z = D[14][j:j + L]
        V6X = D[15][j:j + L]
        V6Y = D[16][j:j + L]
        V6Z = D[17][j:j + L]
        V7X = D[18][j:j + L]
        V7Y = D[19][j:j + L]
        V7Z = D[20][j:j + L]
        V8X = D[21][j:j + L]
        V8Y = D[22][j:j + L]
        V8Z = D[23][j:j + L]


        if Shi(V4X, V4Y, V4Z, V8X, V8Y, V8Z) == 1:

            A = []
            A.append(n)
            n = n + 1

            f1=cot_feture1(V4X, V4Y, V4Z,V7X, V7Y, V7Z,V8X, V8Y, V8Z)
            f2=cot_feture2(V6Y)
            f3,f4=cot_feture3(V5X, V5Y, V5Z,V7X, V7Y, V7Z)
            A.append(f1)
            A.append(f2)
            A.append(f3)
            A.append(f4)

            for k in range (24,29):
                A.append(D[k][j])

            A.append(int(savepath[-6]))


            #print(A)
            sheet.append(A)

    print(n)

    wb.save(savepath)

def shap_data(path):
    wb = openpyxl.Workbook()  # 创建一个excel文件
    sheet = wb.active  # 获得一个的工作表

    sheet = getdata(path, sheet)

    D = []
    n = sheet.max_column
    for i in range(0, n):
        A = []
        D.append(A)

    for row in sheet.iter_rows():

        row_data = [cell.value for cell in row]

        for i in range(0, n):
            # D[i].append(float(row_data[i] - row_data[i % 3]))
            D[i].append(float(row_data[i]))

    L = 10
    # len(D[0])-L
    A=[]
    B=[]

    for j in range(0, len(D[0]) - L + 1):
        # print(j,j+10)

        V4X = D[9][j:j + L]
        V4Y = D[10][j:j + L]
        V4Z = D[11][j:j + L]
        V5X = D[12][j:j + L]
        V5Y = D[13][j:j + L]
        V5Z = D[14][j:j + L]
        V6X = D[15][j:j + L]
        V6Y = D[16][j:j + L]
        V6Z = D[17][j:j + L]
        V7X = D[18][j:j + L]
        V7Y = D[19][j:j + L]
        V7Z = D[20][j:j + L]
        V8X = D[21][j:j + L]
        V8Y = D[22][j:j + L]
        V8Z = D[23][j:j + L]

        if Shi(V4X, V4Y, V4Z, V8X, V8Y, V8Z) == 1:

            E=[]
            F=[]


            f1 = cot_feture1(V4X, V4Y, V4Z, V7X, V7Y, V7Z, V8X, V8Y, V8Z)
            f2 = cot_feture2(V6Y)
            f3, f4 = cot_feture3(V5X, V5Y, V5Z, V7X, V7Y, V7Z)
            E.append(f1)
            E.append(f2)
            E.append(f3)
            E.append(f4)

            for k in range(24, 29):
                F.append(D[k][j])

            A.append(E)
            B.append(F)

    return A,B

if __name__ == '__main__':

    path = './Data/Video/4/'
    savepath1 = './Data/Video/4.xlsx'
    savepath2 = './Excel/Video/4.xlsx'
    videotoexecl(path,savepath1)
    #model_data(savepath1,savepath2)