import openpyxl
import os
import matplotlib

matplotlib.use('TkAgg')
import matplotlib.pyplot as plt
import pandas as pd
import numpy as np
import openpyxl
from openpyxl import load_workbook
import Gate


# 读取execl
def readexecl(path):
    workbook = load_workbook(path, data_only=True)  # 使用data_only=True选项来获取计算后的值而不是公式
    sheet = workbook.active

    D = []
    n = sheet.max_column
    for i in range(n):  # 注意这里从0开始是正确的，但需要确保后续逻辑与之匹配
        A = []
        D.append(A)

    for row in sheet.iter_rows(values_only=True):  # 使用values_only=True来获取单元格的值而非Cell对象
        row_data = list(row)

        for i in range(n):
            cell_value = row_data[i]
            if isinstance(cell_value, (int, float)):  # 确保只对数字类型的值进行转换
                D[i].append(float(cell_value))
            else:
                # 对于非数字类型的值，可以决定如何处理，例如保留原样、替换为None等
                D[i].append(None)  # 或者选择其他处理方式

    return D


def Chu(P1, P2):
    flag = 0
    peaks1 = find_peaks(P1)
    peaks2 = find_peaks(P2)

    if len(peaks1) < len(peaks2):
        flag = 1
    else:
        flag = 0
    return flag


def find_peaks(arr):
    peaks = [arr[0]]  # 初始峰值列表，包含第一个元素
    for i in range(1, len(arr) - 1):
        if arr[i] > arr[i - 1] and arr[i] > arr[i + 1]:
            peaks.append(arr[i])
    peaks.append(arr[-1])  # 添加最后一个元素，如果它是峰值
    return peaks


# 计算向量与坐标轴的夹角
def calculate_angles(vectors):
    """
    计算每个向量与三个坐标轴的夹角 (alpha, beta, gamma)。
    公式：
        alpha = arctan(Ax / sqrt(Ay^2 + Az^2))
        beta  = arctan(Ay / sqrt(Ax^2 + Az^2))
        gamma = arctan(Az / sqrt(Ax^2 + Ay^2))
    """
    angles = []
    for vector in vectors:
        Ax, Ay, Az = vector
        # 避免除以零的情况
        if Ay ** 2 + Az ** 2 == 0:
            alpha = np.nan
        else:
            alpha = np.arctan(Ax / np.sqrt(Ay ** 2 + Az ** 2))

        if Ax ** 2 + Az ** 2 == 0:
            beta = np.nan
        else:
            beta = np.arctan(Ay / np.sqrt(Ax ** 2 + Az ** 2))

        if Ax ** 2 + Ay ** 2 == 0:
            gamma = np.nan
        else:
            gamma = np.arctan(Az / np.sqrt(Ax ** 2 + Ay ** 2))

        angles.append((alpha, beta, gamma))
    return np.array(angles)


# 将弧度转换为角度
def convert_radians_to_degrees(angles):
    """
    将弧度制的角度转换为角度制。
    """
    return np.degrees(angles)


# 新增函数来将一维列表转换为三维向量列表
def convert_to_vectors(data):
    vectors = []
    for i in range(0, len(data), 3):
        if i + 3 <= len(data):  # 确保有足够的元素来形成一个三维向量
            vectors.append(data[i:i + 3])
        else:
            # 处理不足三个元素的情况，例如填充默认值None或跳过
            break  # 或者使用 vectors.append([data[i], None, None]) 来填充
    return vectors


def model_data(execlpath, savepath):
    k=execlpath[-1]

    wb = openpyxl.Workbook()
    sheet = wb.active

    name = ['Num']
    for i in range(1, 10):
        name.append(str(i))
    name.append('Type')

    sheet.append(name)

    n = 0
    file = os.listdir(execlpath)

    for f in file:
            path = os.path.join(execlpath, f)


            D = readexecl(path)

            L = 50
            for j in range(0, len(D[0]) - L):
                P1 = D[0][j:j + L]
                P2 = D[1][j:j + L]
                A1 = D[2][j:j + L]
                A2 = D[3][j:j + L]
                A3 = D[4][j:j + L]

                #P1=Gate.gate(P1,A1)
                #P2=Gate.gate(P2,A2)

                if Chu(P1, P2) == 1:
                    print(path)
                    # 计算P1和P2的相关统计量
                    p1_max = max(P1)
                    p1_avg = np.mean(P1)
                    p1_integral = np.trapz(P1)/L

                    # 转换为三维向量列表
                    A1 = convert_to_vectors(A1)
                    A2 = convert_to_vectors(A2)
                    A3 = convert_to_vectors(A3)

                    A1 = calculate_angles(A1)
                    A1 = convert_radians_to_degrees(A1)

                    A2 = calculate_angles(A2)
                    A2 = convert_radians_to_degrees(A2)

                    A3 = calculate_angles(A3)
                    A3 = convert_radians_to_degrees(A3)

                    # 计算A1, A2, A3的最大值和平均值
                    a1_max = np.max(A1)
                    a1_avg = np.mean(A1)

                    a2_max = np.max(A2)
                    a2_avg = np.mean(A2)

                    a3_max = np.max(A3)
                    a3_avg = np.mean(A3)

                    # 准备输出行B
                    B = []
                    B.append(str(n))
                    n += 1
                    B.extend([p1_max, p1_avg, p1_integral,
                              a1_max, a1_avg, a2_max, a2_avg, a3_max, a3_avg])
                    B.append(k)
                    sheet.append(B)

                   # print(f"P1 Max: {p1_max}, Avg: {p1_avg}, Integral: {p1_integral}")
                   # print(f"A1 Max: {a1_max}, Avg: {a1_avg}")
                   # print(f"A2 Max: {a2_max}, Avg: {a2_avg}")
                   # print(f"A3 Max: {a3_max}, Avg: {a3_avg}")

    wb.save(savepath)


def shap_data(path):

    A = []
    B = []

    D = readexecl(path)

    L = 50
    for j in range(0, len(D[0]) - L):
        P1 = D[0][j:j + L]
        P2 = D[1][j:j + L]
        A1 = D[2][j:j + L]
        A2 = D[3][j:j + L]
        A3 = D[4][j:j + L]


        if Chu(P1, P2) == 1:

            # 计算P1和P2的相关统计量
            p1_max = max(P1)
            p1_avg = np.mean(P1)
            p1_integral = np.trapz(P1) / L

            # 转换为三维向量列表
            A1 = convert_to_vectors(A1)
            A2 = convert_to_vectors(A2)
            A3 = convert_to_vectors(A3)

            A1 = calculate_angles(A1)
            A1 = convert_radians_to_degrees(A1)

            A2 = calculate_angles(A2)
            A2 = convert_radians_to_degrees(A2)

            A3 = calculate_angles(A3)
            A3 = convert_radians_to_degrees(A3)

            # 计算A1, A2, A3的最大值和平均值
            a1_max = np.max(A1)
            a1_avg = np.mean(A1)

            a2_max = np.max(A2)
            a2_avg = np.mean(A2)

            a3_max = np.max(A3)
            a3_avg = np.mean(A3)

            A.append([p1_max, p1_avg, p1_integral])

            B.append([a1_max, a1_avg, a2_max, a2_avg, a3_max, a3_avg])

    return A,B





if __name__ == '__main__':

    savepath1 = './Data/Cot/4'
    savepath2 = './Excel/Cot/4.xlsx'

    model_data(savepath1, savepath2)

















