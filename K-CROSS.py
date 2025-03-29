import openpyxl
from openpyxl import load_workbook
import matplotlib
matplotlib.use('TkAgg')
import matplotlib.pyplot as plt
import numpy as np
import pickle
from sklearn.model_selection import train_test_split
from sklearn.metrics import accuracy_score, recall_score, f1_score, confusion_matrix, ConfusionMatrixDisplay
from sklearn.tree import DecisionTreeClassifier
from sklearn.naive_bayes import GaussianNB
from sklearn.neighbors import KNeighborsClassifier
from sklearn.neural_network import MLPClassifier
from sklearn.ensemble import RandomForestClassifier
from sklearn.svm import SVC
from sklearn.model_selection import KFold


# DecisionTree
def net(path):

    workbook = load_workbook(path)
    sheet = workbook.active

    x_values = []
    y_values = []
    for row in sheet.iter_rows():
        row_data = [cell.value for cell in row]

        x = row_data[1:-1]
        y = row_data[len(row_data) - 1]

        x_values.append(x)
        y_values.append(y)

    features = x_values[0]

    x_values = np.array(x_values[1:])
    y_values = np.array(y_values[1:])

    #P = compare4(x_values, y_values, features)
    #x_values = np.array(bian(x_values, P))

    kf = KFold(n_splits=10)

    U = []
    P=[]
    n=1

    for train_index, test_index in kf.split(x_values):
        x_train, x_test = x_values[train_index], x_values[test_index]
        y_train, y_test = y_values[train_index], y_values[test_index]



        model1 = DecisionTreeClassifier(max_depth=None, min_samples_split=2, random_state=42)


        model=model1


        model.fit(x_train, y_train)




        y_pred = model.predict(x_test)
        #print(y_pred)

        accuracy = accuracy_score(y_test, y_pred)
        print(f"Accuracy: {accuracy:.3f}")
        U.append(round(accuracy, 3))
        P.append(n)
        n=n+1

    m=0
    for i in range(0,len(U)):
        m=m+U[i]
    average=m/10

    plt.rcParams['font.sans-serif'] = ['SimHei']
    plt.rcParams['font.family'] = 'sans-serif'
    plt.rcParams['axes.unicode_minus'] = False

    plt.xlabel('epoch', fontsize=20)
    plt.ylabel('Accuracy', fontsize=20)
    plt.title('10-Fold Cross-Validation  Average-accuracy:'+str(average), fontsize=20)
    plt.ylim(0, 1.1)
    plt.rcParams['font.family'] = 'sans-serif'
    plt.rcParams['font.sans-serif'] = ['Times New Roman']



    plt.plot(P, U, c='blue', linewidth=2, label="accuracy")

    plt.show()

    print(U,P)



# L1-based feature selection
def compare4(x,y,feature_names):
    from sklearn.linear_model import Lasso
    from sklearn.preprocessing import StandardScaler
    scaler = StandardScaler()

    lasso = Lasso(alpha=0.2)
    lasso.fit(x, y)
    lasso.coef_
    #print("Selected features4:", lasso.coef_)

    P = lasso.coef_
    return P


def bian(X,P):
    new_x =[]
    for i in range(0,len(X)):
        x=[]
        for j in range(0, len(P)):
            x.append(X[i][j]*P[j])
        new_x.append(x)
    return new_x


if __name__ == '__main__':


    path9 = './Excel/Gate_Attention_LMF.xlsx'
    path=path9

    net(path)
